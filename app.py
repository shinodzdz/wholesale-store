from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
import os, random, string, hashlib
from datetime import datetime

_basedir = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=os.path.join(_basedir, 'templates'))
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-to-a-random-secret')

db_url = os.environ.get('DATABASE_URL', f'sqlite:///{os.path.join(_basedir, "database.db")}')
db_url = db_url.replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['SHOP_NAME'] = 'Dépot Bouras Béchar'
app.config['SHOP_TAGLINE'] = 'طلب المنتجات الغذائية بالجملة'
app.config['PUBLIC_URL'] = os.environ.get('PUBLIC_URL', '')

db = SQLAlchemy(app)

with app.app_context():
    db.create_all()
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    prod_cols = [c['name'] for c in inspector.get_columns('products')]
    shop_cols = [c['name'] for c in inspector.get_columns('shops')]
    if 'price_semi' not in prod_cols:
        db.session.execute(text('ALTER TABLE products ADD COLUMN price_semi FLOAT DEFAULT 0'))
    if 'type' not in shop_cols:
        db.session.execute(text("ALTER TABLE shops ADD COLUMN type VARCHAR(20) DEFAULT 'تاجر جملة'"))
    db.session.commit()

class Admin(db.Model):
    __tablename__ = 'admins'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)

class Product(db.Model):
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    price = db.Column(db.Float, nullable=False)
    price_semi = db.Column(db.Float, default=0)
    unit = db.Column(db.String(50), default='قطعة')
    category = db.Column(db.String(100), default='عام')
    available = db.Column(db.Integer, default=1)
    stock = db.Column(db.Float, default=0)

class Shop(db.Model):
    __tablename__ = 'shops'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    code = db.Column(db.String(20), unique=True, nullable=False)
    phone = db.Column(db.String(50))
    address = db.Column(db.String(300))
    type = db.Column(db.String(20), default='تاجر جملة')
    created_at = db.Column(db.DateTime, default=datetime.now)

class Order(db.Model):
    __tablename__ = 'orders'
    id = db.Column(db.Integer, primary_key=True)
    shop_id = db.Column(db.Integer, db.ForeignKey('shops.id'), nullable=False)
    shop_name = db.Column(db.String(200))
    notes = db.Column(db.Text)
    total = db.Column(db.Float, default=0)
    status = db.Column(db.String(50), default='جديد')
    created_at = db.Column(db.DateTime, default=datetime.now)
    shop = db.relationship('Shop', backref='orders')

class OrderItem(db.Model):
    __tablename__ = 'order_items'
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('orders.id'), nullable=False)
    product_name = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    price = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(50))

def init_db():
    db.create_all()
    if not Admin.query.first():
        db.session.add(Admin(username='admin', password=hashlib.sha256('admin123'.encode()).hexdigest()))
        db.session.commit()
    try:
        from sqlalchemy import text
        db.session.execute(text('ALTER TABLE products ADD COLUMN stock FLOAT DEFAULT 0'))
        db.session.commit()
    except:
        db.session.rollback()

with app.app_context():
    init_db()

def get_public_url():
    if app.config['PUBLIC_URL']:
        return app.config['PUBLIC_URL'].rstrip('/') + '/'
    try:
        return request.host_url
    except:
        return 'https://'

@app.context_processor
def inject_globals():
    return {
        'shop_name': app.config['SHOP_NAME'],
        'shop_tagline': app.config['SHOP_TAGLINE'],
        'public_url': get_public_url()
    }

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

def generate_code():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

# ─── Admin ───

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = hashlib.sha256(request.form['password'].encode()).hexdigest()
        admin = Admin.query.filter_by(username=username, password=password).first()
        if admin:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            return redirect(url_for('admin'))
        return render_template('admin_login.html', error='خطأ في اسم المستخدم أو كلمة المرور')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin')
@admin_required
def admin():
    shops = Shop.query.order_by(Shop.created_at.desc()).all()
    products = Product.query.filter_by(available=1).order_by(Product.category, Product.name).all()
    new_orders_count = Order.query.filter_by(status='جديد').count()
    pending_count = Order.query.filter(Order.status != 'تم التوصيل').count()
    recent_orders = Order.query.options(db.joinedload(Order.shop)).order_by(Order.created_at.desc()).limit(20).all()
    return render_template('admin.html', shops=shops, products=products, orders=recent_orders,
                           new_orders_count=new_orders_count, pending_count=pending_count)

# ─── Products ───

@app.route('/admin/products', methods=['GET', 'POST'])
@admin_required
def manage_products():
    if request.method == 'POST':
        name = request.form['name']
        price = float(request.form['price'])
        price_semi = float(request.form.get('price_semi', 0) or 0)
        unit = request.form.get('unit', 'قطعة')
        category = request.form.get('category', 'عام')
        stock = float(request.form.get('stock', 0))
        db.session.add(Product(name=name, price=price, price_semi=price_semi, unit=unit, category=category, stock=stock))
        db.session.commit()
    products = Product.query.order_by(Product.category, Product.name).all()
    return render_template('products.html', products=products)

@app.route('/admin/products/delete/<int:id>')
@admin_required
def delete_product(id):
    product = db.session.get(Product, id)
    if product:
        db.session.delete(product)
        db.session.commit()
    return redirect(url_for('manage_products'))

@app.route('/admin/products/bulk_delete', methods=['POST'])
@admin_required
def bulk_delete_products():
    ids = request.form.getlist('ids')
    for pid in ids:
        product = db.session.get(Product, int(pid))
        if product:
            db.session.delete(product)
    db.session.commit()
    return redirect(url_for('manage_products'))

@app.route('/admin/products/toggle/<int:id>')
@admin_required
def toggle_product(id):
    product = db.session.get(Product, id)
    if product:
        product.available = 0 if product.available else 1
        db.session.commit()
    return redirect(url_for('manage_products'))

@app.route('/admin/products/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_product(id):
    product = db.session.get(Product, id)
    if not product:
        return redirect(url_for('manage_products'))
    if request.method == 'POST':
        product.name = request.form['name']
        product.price = float(request.form['price'])
        product.price_semi = float(request.form.get('price_semi', 0) or 0)
        product.unit = request.form.get('unit', 'قطعة')
        product.category = request.form.get('category', 'عام')
        product.stock = float(request.form.get('stock', 0))
        db.session.commit()
        return redirect(url_for('manage_products'))
    return render_template('product_edit.html', product=product)

# ─── Stock Management ───

@app.route('/admin/stock/import', methods=['POST'])
@admin_required
def stock_import():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return redirect(url_for('manage_stock'))
    from openpyxl import load_workbook
    wb = load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return redirect(url_for('manage_stock'))
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    col_map = {}
    for h in headers:
        hl = h.strip()
        if hl in ('الاسم', 'اسم', 'name'):
            col_map['name'] = h
        elif hl in ('المخزون', 'مخزون', 'stock'):
            col_map['stock'] = h
    updated = 0
    for row in rows[1:]:
        vals = [str(v).strip() if v else '' for v in row]
        name = vals[headers.index(col_map['name'])] if 'name' in col_map else ''
        stock = int(float(vals[headers.index(col_map['stock'])])) if 'stock' in col_map and vals[headers.index(col_map['stock'])] else 0
        if name:
            prod = Product.query.filter_by(name=name).first()
            if prod:
                prod.stock = max(0, stock)
                updated += 1
    db.session.commit()
    return redirect(url_for('manage_stock'))

@app.route('/admin/stock', methods=['GET', 'POST'])
@admin_required
def manage_stock():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'bulk_add':
            for key, val in request.form.items():
                if key.startswith('qty_'):
                    qty = int(float(val)) if val else 0
                    if qty > 0:
                        pid = int(key.split('_', 1)[1])
                        prod = db.session.get(Product, pid)
                        if prod:
                            prod.stock = (prod.stock or 0) + qty
            db.session.commit()
            return redirect(url_for('manage_stock'))
        if action == 'bulk_set':
            for key, val in request.form.items():
                if key.startswith('stock_'):
                    qty = int(float(val)) if val else 0
                    pid = int(key.split('_', 1)[1])
                    prod = db.session.get(Product, pid)
                    if prod:
                        prod.stock = max(0, qty)
            db.session.commit()
            return redirect(url_for('manage_stock'))
        product_id = int(request.form.get('product_id'))
        qty = int(float(request.form.get('qty', 0)))
        product = db.session.get(Product, product_id)
        if product:
            if action == 'add':
                product.stock = (product.stock or 0) + qty
            elif action == 'set':
                product.stock = max(0, qty)
            elif action == 'sub':
                product.stock = max(0, (product.stock or 0) - qty)
            db.session.commit()
        return redirect(url_for('manage_stock'))
    products = Product.query.order_by(Product.category, Product.name).all()
    return render_template('stock.html', products=products)

# ─── Shops ───

@app.route('/admin/shops', methods=['GET', 'POST'])
@admin_required
def manage_shops():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form.get('phone', '')
        address = request.form.get('address', '')
        shop_type = request.form.get('type', 'تاجر جملة')
        code = generate_code()
        while Shop.query.filter_by(code=code).first():
            code = generate_code()
        db.session.add(Shop(name=name, code=code, phone=phone, address=address, type=shop_type))
        db.session.commit()
    shops = Shop.query.order_by(Shop.created_at.desc()).all()
    return render_template('shops.html', shops=shops)

@app.route('/admin/shops/delete/<int:id>')
@admin_required
def delete_shop(id):
    shop = db.session.get(Shop, id)
    if shop:
        Order.query.filter_by(shop_id=id).delete()
        db.session.delete(shop)
        db.session.commit()
    return redirect(url_for('manage_shops'))

@app.route('/admin/shops/bulk_delete', methods=['POST'])
@admin_required
def bulk_delete_shops():
    ids = request.form.getlist('ids')
    for sid in ids:
        shop = db.session.get(Shop, int(sid))
        if shop:
            Order.query.filter_by(shop_id=int(sid)).delete()
            db.session.delete(shop)
    db.session.commit()
    return redirect(url_for('manage_shops'))

@app.route('/admin/shops/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_shop(id):
    shop = db.session.get(Shop, id)
    if not shop:
        return redirect(url_for('manage_shops'))
    if request.method == 'POST':
        shop.name = request.form['name']
        shop.phone = request.form.get('phone', '')
        shop.address = request.form.get('address', '')
        shop.type = request.form.get('type', 'تاجر جملة')
        db.session.commit()
        return redirect(url_for('manage_shops'))
    return render_template('shop_edit.html', shop=shop)

# ─── Import Excel ───

@app.route('/admin/import', methods=['GET', 'POST'])
@admin_required
def import_excel():
    result = None
    if request.method == 'POST':
        file = request.files.get('file')
        import_type = request.form.get('type')
        if not file or not file.filename.endswith('.xlsx'):
            return render_template('import.html', error='يرجى رفع ملف Excel بصيغة .xlsx')
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return render_template('import.html', error='الملف لا يحتوي على بيانات كافية')
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            added = 0
            if import_type == 'products':
                col_map = {}
                for h in headers:
                    hl = h.strip()
                    if hl in ('الاسم', 'اسم', 'name'):
                        col_map['name'] = h
                    elif hl in ('سعر الجملة', 'سعرالجملة', 'price'):
                        col_map['price'] = h
                    elif hl in ('سعر التجزئة', 'سعر نصف الجملة', 'سعر النصف جملة', 'سعرنصفالجملة', 'سعرالنصفجملة', 'نصف جملة', 'price_semi', 'semi price'):
                        col_map['price_semi'] = h
                    elif hl in ('الوحدة', 'وحدة', 'unit'):
                        col_map['unit'] = h
                    elif hl in ('التصنيف', 'تصنيف', 'category'):
                        col_map['category'] = h
                    elif hl in ('المخزون', 'مخزون', 'stock'):
                        col_map['stock'] = h
                for row in rows[1:]:
                    vals = [str(v).strip() if v else '' for v in row]
                    name = vals[headers.index(col_map['name'])] if 'name' in col_map else ''
                    if not name:
                        continue
                    price = float(vals[headers.index(col_map['price'])]) if 'price' in col_map and vals[headers.index(col_map['price'])] else 0
                    price_semi = float(vals[headers.index(col_map['price_semi'])]) if 'price_semi' in col_map and vals[headers.index(col_map['price_semi'])] else 0
                    unit = vals[headers.index(col_map['unit'])] if 'unit' in col_map else 'قطعة'
                    category = vals[headers.index(col_map['category'])] if 'category' in col_map else 'عام'
                    stock = int(float(vals[headers.index(col_map['stock'])])) if 'stock' in col_map and vals[headers.index(col_map['stock'])] else 0
                    db.session.add(Product(name=name, price=price, price_semi=price_semi, unit=unit, category=category, stock=stock))
                    added += 1
            elif import_type == 'shops':
                for row in rows[1:]:
                    vals = [str(v).strip() if v else '' for v in row]
                    name = vals[headers.index('name')] if 'name' in headers else (vals[headers.index('الاسم')] if 'الاسم' in headers else (vals[0] if len(vals) > 0 else ''))
                    phone = vals[headers.index('phone')] if 'phone' in headers else (vals[headers.index('الهاتف')] if 'الهاتف' in headers else (vals[1] if len(vals) > 1 else ''))
                    address = vals[headers.index('address')] if 'address' in headers else (vals[headers.index('العنوان')] if 'العنوان' in headers else (vals[2] if len(vals) > 2 else ''))
                    shop_type = vals[headers.index('type')] if 'type' in headers else (vals[headers.index('النوع')] if 'النوع' in headers else 'تاجر جملة')
                    if name:
                        code = generate_code()
                        while Shop.query.filter_by(code=code).first():
                            code = generate_code()
                        db.session.add(Shop(name=name, code=code, phone=phone, address=address, type=shop_type))
                        added += 1
            db.session.commit()
            result = f'✅ تم استيراد {added} {import_type} بنجاح'
        except Exception as e:
            db.session.rollback()
            return render_template('import.html', error=f'خطأ: {str(e)}')
    return render_template('import.html', result=result)

# ─── Shop Orders ───

@app.route('/')
def index():
    return redirect(url_for('shop_login'))

@app.route('/shop', methods=['GET', 'POST'])
def shop_login():
    if request.method == 'POST':
        code = request.form['code'].strip().upper()
        shop = Shop.query.filter_by(code=code).first()
        if shop:
            return redirect(url_for('shop_dashboard', code=code))
        return render_template('shop_login.html', error='كود غير صحيح')
    return render_template('shop_login.html')

@app.route('/shop/<code>')
def shop_dashboard(code):
    shop = Shop.query.filter_by(code=code).first()
    if not shop:
        return redirect(url_for('shop_login'))
    orders = Order.query.filter_by(shop_id=shop.id).order_by(Order.created_at.desc()).limit(10).all()
    return render_template('shop_dashboard.html', shop=shop, orders=orders)

@app.route('/shop/<code>/order')
def shop_order(code):
    shop = Shop.query.filter_by(code=code).first()
    if not shop:
        return redirect(url_for('shop_login'))
    products = Product.query.filter_by(available=1).order_by(Product.category, Product.name).all()
    categories = [row[0] for row in Product.query.filter_by(available=1).with_entities(Product.category).distinct().order_by(Product.category).all()]
    last_order = Order.query.filter_by(shop_id=shop.id).order_by(Order.created_at.desc()).first()
    return render_template('shop_order.html', shop=shop, products=products,
                           categories=categories, last_order=last_order)

@app.route('/shop/<code>/orders')
def shop_orders(code):
    shop = Shop.query.filter_by(code=code).first()
    if not shop:
        return redirect(url_for('shop_login'))
    orders = Order.query.filter_by(shop_id=shop.id).order_by(Order.created_at.desc()).all()
    return render_template('shop_orders.html', shop=shop, orders=orders)

@app.route('/api/submit-order', methods=['POST'])
def submit_order():
    data = request.json
    code = data.get('code')
    items = data.get('items', [])
    notes = data.get('notes', '')

    if not code or not items:
        return jsonify({'success': False, 'error': 'بيانات ناقصة'})

    shop = Shop.query.filter_by(code=code).first()
    if not shop:
        return jsonify({'success': False, 'error': 'كود غير صحيح'})

    total = sum(float(item['price']) * float(item['qty']) for item in items)
    order = Order(shop_id=shop.id, shop_name=shop.name, notes=notes, total=total)
    db.session.add(order)
    db.session.flush()

    for item in items:
        db.session.add(OrderItem(
            order_id=order.id,
            product_name=item['name'],
            quantity=float(item['qty']),
            price=float(item['price']),
            unit=item.get('unit', '')
        ))

    db.session.commit()
    return jsonify({'success': True, 'order_id': order.id})

# ─── View Orders ───

@app.route('/admin/orders')
@admin_required
def view_orders():
    orders = Order.query.options(db.joinedload(Order.shop)).order_by(Order.created_at.desc()).all()
    return render_template('orders.html', orders=orders)

@app.route('/admin/shops/<int:id>/orders')
@admin_required
def shop_orders_admin(id):
    shop = db.session.get(Shop, id)
    if not shop:
        return redirect(url_for('manage_shops'))
    orders = Order.query.options(db.joinedload(Order.shop)).filter_by(shop_id=id).order_by(Order.created_at.desc()).all()
    return render_template('orders.html', orders=orders, shop_filter=shop)

@app.route('/admin/order/<int:id>')
@admin_required
def order_detail(id):
    order = Order.query.options(db.joinedload(Order.shop)).get_or_404(id)
    items = OrderItem.query.filter_by(order_id=id).all()
    return render_template('order_detail.html', order=order, items=items)

@app.route('/admin/order/status/<int:id>', methods=['POST'])
@admin_required
def update_status(id):
    order = db.session.get(Order, id)
    if order:
        old_status = order.status
        order.status = request.form['status']
        if order.status == 'تم التوصيل' and old_status != 'تم التوصيل':
            items = OrderItem.query.filter_by(order_id=id).all()
            for item in items:
                prod = Product.query.filter_by(name=item.product_name).first()
                if prod:
                    prod.stock = max(0, prod.stock - item.quantity)
        db.session.commit()
    return redirect(url_for('view_orders'))

@app.route('/admin/order/delete/<int:id>', methods=['POST'])
@admin_required
def delete_order(id):
    order = db.session.get(Order, id)
    if order:
        OrderItem.query.filter_by(order_id=id).delete()
        db.session.delete(order)
        db.session.commit()
    return redirect(url_for('view_orders'))

@app.route('/admin/change-password', methods=['GET', 'POST'])
@admin_required
def change_password():
    if request.method == 'POST':
        current = hashlib.sha256(request.form['current'].encode()).hexdigest()
        new = request.form['new']
        confirm = request.form['confirm']
        admin = Admin.query.filter_by(username=session['admin_username']).first()
        if admin.password != current:
            return render_template('change_password.html', error='كلمة المرور الحالية غير صحيحة')
        if new != confirm:
            return render_template('change_password.html', error='كلمة المرور الجديدة غير متطابقة')
        if len(new) < 4:
            return render_template('change_password.html', error='كلمة المرور يجب أن تكون 4 أحرف على الأقل')
        admin.password = hashlib.sha256(new.encode()).hexdigest()
        db.session.commit()
        return render_template('change_password.html', success='✅ تم تغيير كلمة المرور بنجاح')
    return render_template('change_password.html')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug)
